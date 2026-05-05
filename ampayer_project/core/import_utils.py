import openpyxl
from datetime import datetime, time
from django.utils.dateparse import parse_date, parse_time
from .models import Team, Category, Player, Stadium, Game, User

def parse_teams_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    results = {'created': 0, 'updated': 0, 'errors': []}
    
    # Expected columns: Name, Category, Manager Name
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name, cat_name, manager = row[0], row[1], row[2]
            
            if not name or not cat_name:
                continue # Skip empty rows
                
            # Find or create category (assuming default reason/league if not strict)
            # Ideally we should pass a season/league context, but for now we look it up by name
            # or create a default one if it doesn't exist.
            # Warning: This might create duplicates if categories share names across seasons.
            # For simplicity, we get the first one or create new.
            category = Category.objects.filter(name=cat_name).first()
            if not category:
                # If no category exists, we can't safely create one without a Season.
                # We'll skip or error.
                # Actually, let's look for ANY active season or just error out.
                # Just error for now.
                results['errors'].append(f"Row {idx}: Category '{cat_name}' not found.")
                continue

            obj, created = Team.objects.update_or_create(
                name=name,
                category=category,
                defaults={'manager_name': manager}
            )
            
            if created:
                results['created'] += 1
            else:
                results['updated'] += 1
                
        except Exception as e:
            results['errors'].append(f"Row {idx}: {str(e)}")
            
    return results

def parse_players_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    results = {'created': 0, 'updated': 0, 'errors': []}
    
    # Expected columns: First Name, Last Name, Jersey Number, Team Name
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            first_name, last_name, jersey, team_name = row[0], row[1], row[2], row[3]
            
            if not first_name or not team_name:
                continue

            team = Team.objects.filter(name=team_name).first()
            if not team:
                results['errors'].append(f"Row {idx}: Team '{team_name}' not found.")
                continue
                
            obj, created = Player.objects.update_or_create(
                first_name=first_name,
                last_name=last_name,
                team=team,
                defaults={'jersey_number': str(jersey) if jersey else ''}
            )
            
            if created:
                results['created'] += 1
            else:
                results['updated'] += 1
                
        except Exception as e:
            results['errors'].append(f"Row {idx}: {str(e)}")
            
    return results

def parse_stadiums_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    results = {'created': 0, 'updated': 0, 'errors': []}
    
    # Expected columns: Name, Address, City, State, Capacity (optional)
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = row[0]
            address = row[1] if len(row) > 1 else ''
            city = row[2] if len(row) > 2 else ''
            state = row[3] if len(row) > 3 else ''
            
            if not name:
                continue

            obj, created = Stadium.objects.update_or_create(
                name=name,
                defaults={
                    'address': address,
                    'city': city,
                    'state': state
                }
            )
            
            if created:
                results['created'] += 1
            else:
                results['updated'] += 1
                
        except Exception as e:
            results['errors'].append(f"Row {idx}: {str(e)}")
            
    return results

def parse_games_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    results = {'created': 0, 'errors': []}
    
    # Expected: Date (YYYY-MM-DD), Time (HH:MM), Local Team, Visitor Team, Stadium, Category
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            date_val, time_val, local_name, visitor_name, stadium_name, cat_name = row[:6]
            
            if not date_val or not local_name or not visitor_name:
                continue
                
            # Parse Date/Time
            if isinstance(date_val, datetime):
                date_obj = date_val.date()
            elif isinstance(date_val, str):
                date_obj = parse_date(date_val)
            else:
                date_obj = None
                
            if isinstance(time_val, time):
                time_obj = time_val
            elif isinstance(time_val, datetime):
                time_obj = time_val.time()
            elif isinstance(time_val, str):
                time_obj = parse_time(time_val)
            else:
                time_obj = None

            if not date_obj or not time_obj:
                results['errors'].append(f"Row {idx}: Invalid date/time format.")
                continue

            # Lookups
            local_team = Team.objects.filter(name=local_name).first()
            visitor_team = Team.objects.filter(name=visitor_name).first()
            stadium = Stadium.objects.filter(name=stadium_name).first()
            category = Category.objects.filter(name=cat_name).first()
            
            if not local_team or not visitor_team:
                results['errors'].append(f"Row {idx}: Teams not found ({local_name}, {visitor_name}).")
                continue
                
            Game.objects.create(
                date=date_obj,
                time=time_obj,
                local_team=local_team,
                visitor_team=visitor_team,
                stadium=stadium,
                category=category,
                title=f"{local_name} vs {visitor_name}"
            )
            results['created'] += 1

        except Exception as e:
            results['errors'].append(f"Row {idx}: {str(e)}")
            
    return results
