import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.db import models
from .models import Game, LineupEntry, Play

def generate_excel_boxscore(game: Game):
    """
    Generate an Excel file with the box score of the game.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Box Score {game.id}"

    # Header
    ws['A1'] = f"{game.local_team.name} vs {game.visitor_team.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:I1')
    
    ws['A2'] = f"Estadio: {game.stadium.name} | Fecha: {game.date}"
    ws.merge_cells('A2:I2')

    row_num = 4

    # Add Line Score (simplified representation for Excel)
    ws.cell(row=row_num, column=1, value="LINE SCORE").font = Font(bold=True)
    row_num += 1
    
    # Inning headers
    innings_count = 9 # Standard
    ws.cell(row=row_num, column=1, value="Equipo")
    for i in range(1, innings_count + 1):
        ws.cell(row=row_num, column=i+1, value=i).font = Font(bold=True)
    ws.cell(row=row_num, column=innings_count+2, value="R").font = Font(bold=True)
    row_num += 1
    
    # Visitor
    ws.cell(row=row_num, column=1, value=game.visitor_team.name)
    ws.cell(row=row_num, column=innings_count+2, value=game.away_score)
    row_num += 1
    # Local
    ws.cell(row=row_num, column=1, value=game.local_team.name)
    ws.cell(row=row_num, column=innings_count+2, value=game.home_score)
    row_num += 2

    # Function to add team stats
    def add_team_stats(team, lineups, title):
        nonlocal row_num
        ws.cell(row=row_num, column=1, value=title).font = Font(bold=True)
        row_num += 1
        
        headers = ['Jugador', 'Pos', 'AB', 'R', 'H', 'RBI', 'BB', 'SO', 'AVG']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
        row_num += 1

        for entry in lineups:
            avg = entry.calculate_avg()
            ws.cell(row=row_num, column=1, value=f"{entry.player.first_name} {entry.player.last_name}")
            ws.cell(row=row_num, column=2, value=entry.field_position)
            ws.cell(row=row_num, column=3, value=entry.AB)
            ws.cell(row=row_num, column=4, value=entry.R)
            ws.cell(row=row_num, column=5, value=entry.H)
            ws.cell(row=row_num, column=6, value=entry.RBI)
            ws.cell(row=row_num, column=7, value=entry.BB)
            ws.cell(row=row_num, column=8, value=entry.SO)
            ws.cell(row=row_num, column=9, value=str(avg))
            row_num += 1
        
        row_num += 1
        
        # Pitching header
        ws.cell(row=row_num, column=1, value="Pitching").font = Font(bold=True)
        row_num += 1
        p_headers = ['Pitcher', 'IP', 'H', 'R', 'ER', 'BB', 'SO', 'HR']
        for col, header in enumerate(p_headers, 1):
            ws.cell(row=row_num, column=col, value=header).font = Font(bold=True)
        row_num += 1
        
        pitchers = lineups.filter(IP_outs__gt=0) | lineups.filter(field_position='1')
        for p in pitchers.distinct():
            ip = f"{p.IP_outs // 3}.{p.IP_outs % 3}"
            ws.cell(row=row_num, column=1, value=f"{p.player.first_name} {p.player.last_name}")
            ws.cell(row=row_num, column=2, value=ip)
            ws.cell(row=row_num, column=3, value=p.pitch_H)
            ws.cell(row=row_num, column=4, value=p.pitch_R)
            ws.cell(row=row_num, column=5, value=p.pitch_ER)
            ws.cell(row=row_num, column=6, value=p.pitch_BB)
            ws.cell(row=row_num, column=7, value=p.pitch_SO)
            ws.cell(row=row_num, column=8, value=p.pitch_HR)
            row_num += 1

        row_num += 2

    # Local Team
    local_lineup = LineupEntry.objects.filter(game=game, team=game.local_team).order_by('batting_order')
    add_team_stats(game.local_team, local_lineup, f"Equipo Local: {game.local_team.name}")

    # Visitor Team
    visitor_lineup = LineupEntry.objects.filter(game=game, team=game.visitor_team).order_by('batting_order')
    add_team_stats(game.visitor_team, visitor_lineup, f"Equipo Visitante: {game.visitor_team.name}")

    # Adjust widths
    ws.column_dimensions['A'].width = 25
    

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=boxscore_{game.id}.xlsx'
    wb.save(response)
    return response

def generate_pdf_boxscore(game: Game):
    """
    Generate a PDF file with the box score.
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=boxscore_{game.id}.pdf'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']

    # Header
    elements.append(Paragraph(f"{game.local_team.name} vs {game.visitor_team.name}", title_style))
    elements.append(Paragraph(f"Estadio: {game.stadium.name} | Fecha: {game.date}", normal_style))
    elements.append(Spacer(1, 12))

    # Helper for table data
    def get_team_data(team, lineups):
        data = [['Jugador', 'Pos', 'AB', 'R', 'H', 'RBI', 'BB', 'SO']]
        for entry in lineups:
            data.append([
                f"{entry.player.first_name} {entry.player.last_name}",
                entry.field_position,
                str(entry.AB),
                str(entry.R),
                str(entry.H),
                str(entry.RBI),
                str(entry.BB),
                str(entry.SO),
            ])
        return data

    # Local Team Table
    elements.append(Paragraph(f"<b>Equipo Local: {game.local_team.name}</b>", normal_style))
    local_lineup = LineupEntry.objects.filter(game=game, team=game.local_team).order_by('batting_order')
    local_data = get_team_data(game.local_team, local_lineup)
    
    t_local = Table(local_data)
    t_local.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(t_local)
    elements.append(Spacer(1, 20))

    # Visitor Team Table
    elements.append(Paragraph(f"<b>Equipo Visitante: {game.visitor_team.name}</b>", normal_style))
    visitor_lineup = LineupEntry.objects.filter(game=game, team=game.visitor_team).order_by('batting_order')
    visitor_data = get_team_data(game.visitor_team, visitor_lineup)
    
    t_visitor = Table(visitor_data)
    t_visitor.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(t_visitor)

    doc.build(elements)
    return response

def generate_digital_acta(game: Game):
    """
    Generate a comprehensive Digital Minute (Acta Digital) PDF.
    Includes Box Score, Substitutions, and Play-by-Play.
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=acta_digital_{game.id}.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = styles['Title']
    subheader_style = styles['Heading2']
    section_style = styles['Heading3']
    normal_style = styles['Normal']

    # 1. Page Header
    elements.append(Paragraph(f"ACTA DIGITAL DE JUEGO", header_style))
    elements.append(Paragraph(f"<b>{game.local_team.name} vs {game.visitor_team.name}</b>", subheader_style))
    elements.append(Spacer(1, 10))
    
    # 2. Game Info
    info_data = [
        [f"Fecha: {game.date}", f"Hora: {game.time}", f"Estadio: {game.stadium.name}"],
        [f"Categoría: {game.category.name if game.category else '-'}", f"Temporada: {game.season.name if game.season else '-'}", ""],
        [f"Anotador: {game.scorer_1 if game.scorer_1 else '-'}", f"Ampayer: {game.ampayer_1 if game.ampayer_1 else '-'}", ""]
    ]
    t_info = Table(info_data, colWidths=[180, 150, 180])
    t_info.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
    ]))
    elements.append(t_info)
    elements.append(Spacer(1, 15))

    # 3. Line Score (Carreras por Entrada)
    elements.append(Paragraph("PUNTUACIÓN POR ENTRADAS", section_style))
    innings_count = max(1, game.current_inning)
    line_score_headers = ['EQUIPO'] + [str(i) for i in range(1, innings_count + 1)] + ['R', 'H', 'E']
    
    def get_runs_for_team(side):
        half = 'top' if side == 'visitor' else 'bottom'
        runs = []
        for i in range(1, innings_count + 1):
            inn_runs = game.plays.filter(inning=i, half=half).aggregate(total=models.Sum('runs_scored'))['total'] or 0
            runs.append(str(inn_runs) if i <= game.current_inning else "-")
        return runs

    visitor_runs = get_runs_for_team('visitor')
    local_runs = get_runs_for_team('local')
    
    ls_data = [
        line_score_headers,
        [game.visitor_team.name] + visitor_runs + [str(game.away_score), "-", "-"],
        [game.local_team.name] + local_runs + [str(game.home_score), "-", "-"]
    ]
    
    t_ls = Table(ls_data, hAlign='LEFT')
    t_ls.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 8)
    ]))
    elements.append(t_ls)
    elements.append(Spacer(1, 15))

    # 4. Lineups and Pitching
    def add_team_section(title, team):
        elements.append(Paragraph(title, section_style))
        lineup = LineupEntry.objects.filter(game=game, team=team).order_by('batting_order', 'entry_inning')
        
        # Batting Table
        data = [['#', 'Jugador', 'Pos', 'E. Ent', 'E. Sal', 'AB', 'R', 'H', 'RBI', 'BB', 'SO']]
        for entry in lineup:
            data.append([
                str(entry.batting_order),
                f"{entry.player.first_name} {entry.player.last_name}",
                entry.field_position,
                str(entry.entry_inning),
                str(entry.exit_inning) if entry.exit_inning else "-",
                str(entry.AB), str(entry.R), str(entry.H), str(entry.RBI), str(entry.BB), str(entry.SO)
            ])
        
        t = Table(data, hAlign='LEFT')
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e3a8a")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 8)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 10))

        # Pitching Table for this team
        elements.append(Paragraph(f"Estadísticas de Pitcheo: {team.name}", ParagraphStyle('small', fontSize=8, leading=10, textColor=colors.grey)))
        pitchers = lineup.filter(models.Q(IP_outs__gt=0) | models.Q(field_position='1'))
        if pitchers.exists():
            p_data = [['Pitcher', 'IP', 'H', 'R', 'ER', 'BB', 'SO', 'HR']]
            for p in pitchers.distinct():
                ip = f"{p.IP_outs // 3}.{p.IP_outs % 3}"
                p_data.append([
                    f"{p.player.first_name} {p.player.last_name}",
                    ip, str(p.pitch_H), str(p.pitch_R), str(p.pitch_ER), str(p.pitch_BB), str(p.pitch_SO), str(p.pitch_HR)
                ])
            
            tp = Table(p_data, hAlign='LEFT')
            tp.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#374151")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('FONTSIZE', (0,0), (-1,-1), 7)
            ]))
            elements.append(tp)
            elements.append(Spacer(1, 10))

    add_team_section(f"Equipo Local: {game.local_team.name}", game.local_team)
    add_team_section(f"Equipo Visitante: {game.visitor_team.name}", game.visitor_team)

    # 4. Play-by-Play Log
    elements.append(Paragraph("Resumen de Jugadas / Bitácora", section_style))
    plays = game.plays.all().order_by('id')
    
    play_data = [['Inn', 'Mitad', 'Bateador', 'Evento', 'R', 'O', 'Detalle']]
    for p in plays:
        play_data.append([
            str(p.inning),
            "Alta" if p.half == 'top' else "Baja",
            f"{p.batter.first_name if p.batter else 'Sust'}",
            p.event_type.capitalize(),
            str(p.runs_scored),
            str(p.outs_recorded),
            p.play_description or "-"
        ])
    
    if len(play_data) > 1:
        t_plays = Table(play_data, colWidths=[30, 40, 90, 80, 20, 20, 200])
        t_plays.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#d1d5db")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 7)
        ]))
        elements.append(t_plays)
    else:
        elements.append(Paragraph("No hay jugadas registradas aún.", normal_style))

    # 5. Summary and Signatures
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"Puntuación Final: {game.local_team.name} {game.home_score} - {game.away_score} {game.visitor_team.name}", normal_style))
    
    sig_data = [
        ["_"*30, "_"*30],
        ["Firma Anotador", "Firma Ampayer Principal"]
    ]
    t_sig = Table(sig_data, colWidths=[250, 250])
    t_sig.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 40)
    ]))
    elements.append(t_sig)

    doc.build(elements)
    return response


def generate_umpire_distribution_pdf(games_queryset):
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import HRFlowable
    import datetime

    response = HttpResponse(content_type="application/pdf")
    today_str = str(datetime.date.today())
    response["Content-Disposition"] = "attachment; filename=distribucion_ampayers_" + today_str + ".pdf"

    doc = SimpleDocTemplate(
        response, pagesize=letter,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    elements = []
    styles = getSampleStyleSheet()

    C_NAVY    = colors.HexColor("#1e3a5f")
    C_BLUE    = colors.HexColor("#2563eb")
    C_LIGHT   = colors.HexColor("#e8f0fe")
    C_WHITE   = colors.white
    C_GRAY50  = colors.HexColor("#f9fafb")
    C_GRAY200 = colors.HexColor("#e5e7eb")
    C_GRAY700 = colors.HexColor("#374151")
    C_GREEN   = colors.HexColor("#16a34a")
    C_AMBER   = colors.HexColor("#d97706")
    C_RED     = colors.HexColor("#dc2626")

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_s    = ps("RT",  fontSize=18, fontName="Helvetica-Bold", textColor=C_WHITE, alignment=TA_CENTER)
    subtitle_s = ps("RS",  fontSize=9,  fontName="Helvetica",      textColor=C_LIGHT, alignment=TA_CENTER)
    head_s     = ps("HS",  fontSize=7,  fontName="Helvetica-Bold", textColor=C_WHITE, alignment=TA_CENTER)
    cell_s     = ps("CS",  fontSize=7,  fontName="Helvetica",      textColor=C_GRAY700, alignment=TA_CENTER)
    bold_s     = ps("BS",  fontSize=7,  fontName="Helvetica-Bold", textColor=C_NAVY, alignment=TA_CENTER)
    footer_s   = ps("FS",  fontSize=7,  fontName="Helvetica",      textColor=colors.HexColor("#9ca3af"), alignment=TA_CENTER)
    stat_s     = ps("STC", fontSize=8,  fontName="Helvetica-Bold", textColor=C_WHITE, alignment=TA_CENTER)

    hdr_t = Table([[Paragraph("DISTRIBUCION DE AMPAYERS Y ANOTADORES", title_s)]], colWidths=[doc.width])
    hdr_t.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), C_NAVY),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0),(-1,-1), 14),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
    ]))
    elements.append(hdr_t)

    date_label = datetime.date.today().strftime("%d/%m/%Y")
    sub_t = Table([[Paragraph("Coordinacion de Oficiales  |  Emitido: " + date_label, subtitle_s)]], colWidths=[doc.width])
    sub_t.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), C_BLUE),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    elements.append(sub_t)
    elements.append(Spacer(1, 14))

    games_list = list(games_queryset.select_related(
        "local_team","visitor_team","stadium","category",
        "ampayer_1","ampayer_2","ampayer_3","scorer_1","scorer_2"
    ).order_by("date","time"))

    total      = len(games_list)
    assigned   = sum(1 for g in games_list if g.ampayer_1)
    unassigned = total - assigned

    stats_t = Table([[
        Paragraph("<b>" + str(total) + "</b>\nJuegos Totales", stat_s),
        Paragraph("<b>" + str(assigned) + "</b>\nCon Ampayer",  stat_s),
        Paragraph("<b>" + str(unassigned) + "</b>\nPor Asignar", stat_s),
    ]], colWidths=[doc.width/3]*3)
    stats_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(0,-1), C_BLUE),
        ("BACKGROUND", (1,0),(1,-1), C_GREEN),
        ("BACKGROUND", (2,0),(2,-1), C_AMBER if unassigned > 0 else C_GREEN),
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0),(-1,-1), 10),
        ("BOTTOMPADDING",(0,0),(-1,-1), 10),
        ("GRID",       (0,0),(-1,-1), 1, C_WHITE),
    ]))
    elements.append(stats_t)
    elements.append(Spacer(1, 12))

    col_widths = [22, 65, 42, 38, 105, 80, 95, 78]

    def amp_cell(game):
        lines = []
        labels = ["Principal","Base 1","Base 2"]
        for i, field in enumerate(["ampayer_1","ampayer_2","ampayer_3"]):
            amp = getattr(game, field, None)
            if amp:
                lines.append(str(i+1) + ". " + amp.first_name + " " + amp.last_name + " (" + labels[i] + ")")
        if not lines:
            return Paragraph("<i>Sin asignar</i>",
                ps("na" + str(id(game)), fontSize=7, textColor=C_RED, fontName="Helvetica-Oblique"))
        return Paragraph("<br/>".join(lines),
            ps("ac" + str(id(game)), fontSize=7, fontName="Helvetica", textColor=C_NAVY))

    def scr_cell(game):
        lines = []
        if game.scorer_1:
            lines.append("Ofic. " + game.scorer_1.first_name + " " + game.scorer_1.last_name)
        if game.scorer_2:
            lines.append("Aux. " + game.scorer_2.first_name + " " + game.scorer_2.last_name)
        if not lines:
            return Paragraph("<i>Sin asignar</i>",
                ps("na2" + str(id(game)), fontSize=7, textColor=colors.HexColor("#9ca3af"), fontName="Helvetica-Oblique"))
        return Paragraph("<br/>".join(lines),
            ps("sc" + str(id(game)), fontSize=7, fontName="Helvetica", textColor=C_GRAY700))

    headers = ["#","Categoria","Fecha","Hora","Juego","Estadio","Ampayers","Anotadores"]
    table_data = [[Paragraph("<b>" + h + "</b>", head_s) for h in headers]]

    for idx, game in enumerate(games_list, 1):
        match_s = ps("mt" + str(idx), fontSize=7, fontName="Helvetica-Bold", textColor=C_NAVY, alignment=TA_CENTER)
        row = [
            Paragraph("<b>" + str(idx) + "</b>", bold_s),
            Paragraph(game.category.name if game.category else "-", cell_s),
            Paragraph(game.date.strftime("%d/%m/%y") if game.date else "-", cell_s),
            Paragraph(str(game.time)[:5] if game.time else "-", cell_s),
            Paragraph("<b>" + game.local_team.name + "</b><br/>vs<br/><b>" + game.visitor_team.name + "</b>", match_s),
            Paragraph(game.stadium.name if game.stadium else "-", cell_s),
            amp_cell(game),
            scr_cell(game),
        ]
        table_data.append(row)

    main_t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t_cmds = [
        ("BACKGROUND",    (0,0),(-1,0),  C_NAVY),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("FONTSIZE",      (0,0),(-1,-1), 7),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING",   (0,0),(-1,-1), 3),
        ("RIGHTPADDING",  (0,0),(-1,-1), 3),
        ("GRID",          (0,0),(-1,-1), 0.5, C_GRAY200),
        ("LINEBELOW",     (0,0),(-1,0),  1.5, C_BLUE),
    ]
    for i in range(1, len(table_data)):
        t_cmds.append(("BACKGROUND", (0,i),(-1,i), C_GRAY50 if i % 2 == 0 else C_WHITE))
    main_t.setStyle(TableStyle(t_cmds))
    elements.append(main_t)

    umpire_load = {}
    for game in games_list:
        for field in ["ampayer_1","ampayer_2","ampayer_3"]:
            amp = getattr(game, field, None)
            if amp:
                key = amp.first_name + " " + amp.last_name
                entry = (game.local_team.name + " vs " + game.visitor_team.name +
                         " (" + (game.date.strftime("%d/%m") if game.date else "") + ")")
                umpire_load.setdefault(key, []).append(entry)

    if umpire_load:
        elements.append(Spacer(1, 18))
        elements.append(HRFlowable(width=doc.width, thickness=1.5, color=C_NAVY))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph("RESUMEN DE CARGA POR AMPAYER",
            ps("rh", fontSize=10, fontName="Helvetica-Bold", textColor=C_NAVY, spaceAfter=6)))

        sum_col_w = [130, 45, doc.width - 175]
        sum_data = [[
            Paragraph("<b>Ampayer</b>", head_s),
            Paragraph("<b>Juegos</b>",  head_s),
            Paragraph("<b>Partidos Asignados</b>", head_s),
        ]]
        for i, (name, glist) in enumerate(sorted(umpire_load.items(), key=lambda x: -len(x[1]))):
            sum_data.append([
                Paragraph("<b>" + name + "</b>", ps("sn" + str(i), fontSize=8, fontName="Helvetica-Bold", textColor=C_NAVY)),
                Paragraph("<b>" + str(len(glist)) + "</b>", ps("sc2" + str(i), fontSize=9, fontName="Helvetica-Bold",
                    textColor=C_BLUE, alignment=TA_CENTER)),
                Paragraph("  .  ".join(glist), ps("sd" + str(i), fontSize=7, textColor=C_GRAY700)),
            ])

        sum_t = Table(sum_data, colWidths=sum_col_w)
        sc = [
            ("BACKGROUND",    (0,0),(-1,0),  C_NAVY),
            ("GRID",          (0,0),(-1,-1), 0.5, C_GRAY200),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ("LEFTPADDING",   (0,0),(-1,-1), 4),
        ]
        for i in range(1, len(sum_data)):
            sc.append(("BACKGROUND", (0,i),(-1,i), C_LIGHT if i % 2 == 0 else C_WHITE))
        sum_t.setStyle(TableStyle(sc))
        elements.append(sum_t)

    elements.append(Spacer(1, 18))
    elements.append(HRFlowable(width=doc.width, thickness=0.5, color=C_GRAY200))
    elements.append(Spacer(1, 6))
    date_gen = datetime.date.today().strftime("%d/%m/%Y")
    sig_t = Table([[
        Paragraph("____________________________\nFirma Coordinador de Ampayers", footer_s),
        Paragraph("Sistema Ampayer Profe Bernal\nGenerado el " + date_gen, footer_s),
        Paragraph("____________________________\nFirma Presidente de Liga", footer_s),
    ]], colWidths=[doc.width/3]*3)
    sig_t.setStyle(TableStyle([
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
        ("VALIGN",     (0,0),(-1,-1), "TOP"),
        ("TOPPADDING", (0,0),(-1,-1), 8),
    ]))
    elements.append(sig_t)

    doc.build(elements)
    return response
